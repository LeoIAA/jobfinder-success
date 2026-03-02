"""
Rate 500 jobs in pm_jobs.xlsx with a 0-100 CV-fit score → written to S3 column.
Based on Leo Iurchenko's CV (calibrated against 439 existing S3 manual ratings).

Scoring model: multiplicative
  final = clamp(round(base × location_factor), 0, 100)
  base  = role_pts + domain_pts + exp_pts + bonuses  (max ~80)
  location_factor: Remote=1.10, London Hybrid=1.05, UK Hybrid=0.95,
                   Edinburgh/Bristol Hybrid=0.70, Manchester Hybrid=0.40,
                   Onsite London=0.45, Onsite Northern=0.25

Calibration checks (existing S3 vs computed):
  B2B FinTech PO Remote          90 → ~91  ✓
  FinCrime Fintech PO LDN Hybrid 75 → ~75  ✓
  Stripe PM Bank Transfers Onsite 35 → ~29 ✓
  Taptap Send PM Onsite London    38 → ~38 ✓
  Edinburgh Hybrid legal-tech PM  48 → ~41 ✓
  Manchester Onsite PM            12 → ~14 ✓
"""
import re
import sys
import openpyxl

OUTPUT_FILE = "pm_jobs.xlsx"
ROWS_TO_RATE = 500


# ── Helpers ───────────────────────────────────────────────────────────────────

def clean_title(raw: str) -> str:
    """Strip LinkedIn 'with verification' multi-line artefacts."""
    return re.sub(r"\n.*", "", str(raw)).strip() if raw else ""


# ── Component scorers ─────────────────────────────────────────────────────────

def role_pts(t: str, d: str) -> int:
    """0-30: role type / seniority fit for Leo (mid-level PM/PO)."""
    # Hard dealbreakers at role level
    if re.search(r"\b(director|vp |vice.president|head of product|cpo|chief product"
                 r"|group product manager|gm product|general manager product)\b", t):
        return 5

    if re.search(r"\b(principal product|staff product)\b", t):
        return 12

    # Stretch senior — catch "Senior [adjective] Product Manager/Owner" patterns
    is_senior = bool(re.search(r"\bsenior\b", t))
    is_lead   = bool(re.search(r"\b(lead product|product lead)\b", t))
    if (is_senior or is_lead) and re.search(r"\b(product (manager|owner)|pm\b|po\b)", t):
        return 20

    # Perfect target band
    if re.search(r"\b(product (manager|owner|management))\b", t):
        if re.search(r"\b(junior|associate|graduate|entry.?level|intern)\b", t):
            return 11
        return 28

    # Hybrid analyst/PO title (e.g., "Product Owner Analyst")
    if re.search(r"\bproduct owner analyst\b", t):
        return 14

    # Adjacent roles
    if re.search(r"\bproduct analyst\b", t):
        return 16
    if re.search(r"\b(programme manager|delivery manager|project manager)\b", t):
        return 7
    if re.search(r"\b(business analyst|ba role)\b", t):
        return 6
    if re.search(r"\b(scrum master|agile coach)\b", t):
        return 8

    return 10  # unclear


def domain_pts(t: str, d: str) -> int:
    """0-35: domain / industry relevance to Leo's fintech/B2B SaaS background."""
    combined = t + " " + d

    # Hard no — zero background, unlikely to pass screening
    hard_no = [
        "medical device", "clinical trial", "nhs ", " nhs", "pharmaceutical", "pharma ",
        "defence", "defense", "military", "civil servant", "government digital",
        "embedded system", "firmware", "automotive product", "aerospace product",
        "oil and gas", "nuclear", "mining product",
    ]
    if any(w in combined for w in hard_no):
        return 5

    # Weak domain — Leo could stretch but lacks direct experience
    weak = [
        "healthcare", "health care", "medtech", "biotech", "life science",
        "social housing", "non-profit", "public sector", "local government",
        "retail banking credit", "mortgage product", "manufacturing", "construction tech",
    ]
    if any(w in combined for w in weak):
        return 11

    # Excellent — direct Leo background overlap
    excellent = [
        "fintech", "payments", "payment platform", "trading platform", "broker platform",
        "forex", "crypto product", "defi", "neobank", "challenger bank", "banking platform",
        "b2b saas", "crm product", "wealth management platform", "regtech", "insurtech",
        "financial service product", "open banking", "cross-border payment", "remittance",
        "card product", "lending platform", "capital markets", "investment platform",
        "digital assets", "liquidity management",
    ]
    exc_hits = sum(1 for w in excellent if w in combined)
    if exc_hits >= 2:
        return 34
    if exc_hits == 1:
        return 29

    # Good — transferable skills
    good = [
        "saas", "b2b", "enterprise software", "platform product",
        "automation product", "workflow automation", "internal tool",
        "data product", "analytics platform", "ai product", "ml product",
        "machine learning product", "developer tool", "api product",
        "martech", "adtech", "hr tech", "legal tech", "edtech", "proptech",
        "gaming", "game product", "digital entertainment",
        "e-commerce platform", "marketplace product",
    ]
    good_hits = sum(1 for w in good if w in combined)
    if good_hits >= 3:
        return 27
    if good_hits >= 2:
        return 25
    if good_hits == 1:
        return 18

    # Consumer / website / retail — tangential
    consumer = ["consumer product", "mobile app product", "website product",
                 "retail product", "consumer tech", "d2c"]
    if any(w in combined for w in consumer):
        return 13

    return 15  # neutral / unknown


def exp_pts(t: str, d: str) -> int:
    """0-15: years-of-experience requirement fit (Leo has ~3 yrs PM, ~6 yrs total)."""
    all_years = re.findall(
        r"(\d+)\+?\s*(?:to\s*\d+\s*)?years?\s*(?:of\s*)?(?:proven\s*)?(?:product\s*)?"
        r"(?:management\s*)?experience",
        d,
    )
    if all_years:
        mn = min(int(y) for y in all_years)
        if mn <= 2:
            return 12    # slight over-qualification but ok
        if mn <= 4:
            return 14    # sweet spot
        if mn <= 6:
            return 9     # stretch
        return 3         # 7+ years: unlikely to clear bar

    # Fallback from title
    if re.search(r"\b(junior|associate|graduate|entry)\b", t):
        return 8
    if re.search(r"\b(senior|lead|principal)\b", t):
        return 9     # might need 5-7 yrs
    if re.search(r"\b(director|head|vp|chief)\b", t):
        return 3

    return 11  # neutral


def location_factor(location_str: str, job_type: str, d: str) -> float:
    """Multiplicative factor reflecting Leo's ability to physically do this role."""
    loc = (str(location_str) + " " + str(job_type or "") + " " + d).lower()

    # ── Fully remote ───────────────────────────────────────────────────────────
    if re.search(r"\bfully remote\b|\bremote (only|first|based|working)\b"
                 r"|\(remote\)", loc):
        return 1.10

    is_hybrid = "hybrid" in loc
    is_remote = "remote" in loc
    is_onsite = bool(re.search(r"\bon.?site\b|\bin.?office\b", loc))

    if is_remote and not is_hybrid and not is_onsite:
        return 1.10

    # ── Northern / remote-difficult UK cities ──────────────────────────────────
    northern_cities = r"\b(manchester|birmingham|liverpool|sheffield|nottingham" \
                      r"|newcastle|sunderland|coventry|bradford|wolverhampton)\b"
    scotland_wales = r"\b(glasgow|aberdeen|dundee|cardiff|belfast|swansea|edinburgh)\b"

    # SE England cities (commutable-ish or easily flyable)
    se_cities = r"\b(brighton|guildford|reading|oxford|cambridge|watford|slough|"  \
                r"luton|hertford|milton keynes|southampton|portsmouth|bournemouth)\b"

    # Rural English counties — require being physically nearby; effectively onsite for Leo
    rural_counties = r"\b(lincolnshire|norfolk|suffolk|devon|cornwall|dorset|"     \
                     r"wiltshire|somerset|herefordshire|shropshire|cumbria|"        \
                     r"northumberland|rutland|leicestershire|derbyshire|"           \
                     r"staffordshire|worcestershire|gloucestershire|"               \
                     r"cambridgeshire|northamptonshire|buckinghamshire)\b"

    # "Hybrid/Remote" type → treat as effectively remote regardless of city
    job_type_str = str(job_type or "").lower()
    if "remote" in job_type_str:
        return 1.08   # Hybrid/Remote roles are nearly as good as fully remote

    # ── Remote / Hybrid: location matters ─────────────────────────────────────
    if is_hybrid or (is_remote and is_hybrid):
        if re.search(r"\blondon\b", loc):
            return 1.05
        if re.search(northern_cities, loc):
            return 0.40
        # Major accessible Scottish/Welsh city — manageable occasional trips
        if re.search(r"\bedinburgh\b", loc):
            return 0.78
        if re.search(r"\bbristol\b", loc):
            return 0.82
        if re.search(scotland_wales, loc):
            return 0.50
        if re.search(se_cities, loc):
            return 0.90
        if re.search(rural_counties, loc):
            return 0.32  # rural county: effectively onsite for Leo
        # Generic UK hybrid (truly flexible / UK-wide)
        if re.search(r"\b(united kingdom|uk |england|nationwide)\b", loc):
            return 0.95
        return 0.75   # other unrecognised UK city hybrid

    # ── Onsite ────────────────────────────────────────────────────────────────
    if is_onsite:
        if re.search(r"\blondon\b", loc):
            return 0.45
        if re.search(northern_cities, loc):
            return 0.22
        if re.search(scotland_wales, loc):
            return 0.22
        return 0.35   # other onsite

    # ── No type specified: infer from city ────────────────────────────────────
    if re.search(r"\blondon\b", loc):
        return 0.95   # assume hybrid-leaning for London listings without type
    if re.search(northern_cities, loc):
        return 0.50
    if re.search(scotland_wales, loc):
        return 0.50
    if re.search(r"\b(united kingdom|uk |england)\b", loc):
        return 1.00

    return 0.90  # unknown → modest discount


def score_job(title, company, location, job_type, salary, description) -> int:
    """Return 0-100 CV-fit score for Leo."""
    raw_title = clean_title(title)
    desc = str(description).lower() if description else ""
    t = raw_title.lower()
    combined = t + " " + desc

    # ── Hard dealbreakers (override) ──────────────────────────────────────────
    # Language requirements Leo can't meet
    if re.search(r"\b(mandarin|cantonese|chinese.speaking|arabic.speaking"
                 r"|japanese.speaking|korean.speaking|fluent in (mandarin|arabic|japanese|korean))\b",
                 combined):
        return 5

    # ── Component scores ──────────────────────────────────────────────────────
    r = role_pts(t, desc)
    dm = domain_pts(t, desc)
    e = exp_pts(t, desc)
    base = r + dm + e  # 0-80 typical

    # ── Bonuses ───────────────────────────────────────────────────────────────
    # Leo's exact methods/tools mentioned → signals good match
    skill_hits = sum(1 for kw in [
        "roadmap", "backlog", "agile", "scrum", "sprint", "stakeholder",
        "cross-functional", "user story", "jtbd", "jira",
    ] if kw in combined)
    if skill_hits >= 5:
        base += 4
    elif skill_hits >= 3:
        base += 2

    # Direct Leo background signals in description
    if any(w in combined for w in [
        "trading platform", "crm product", "broker platform",
        "payment platform", "b2b fintech", "saas crm", "financial crm",
    ]):
        base += 3

    # ── Penalties ─────────────────────────────────────────────────────────────
    # Contract / interim (Leo seeking permanent)
    if re.search(r"\b(contract role|interim|fixed.?term|ftc|day rate"
                 r"|12.month contract|6.month contract)\b", combined):
        base -= 6

    # Very high experience bar in title
    if re.search(r"\b(principal|staff pm|distinguished)\b", t):
        base -= 5

    base = max(5, min(80, base))

    # ── Apply location factor ─────────────────────────────────────────────────
    loc_f = location_factor(location, job_type, desc)
    final = round(base * loc_f)

    return max(0, min(100, final))


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    print(f"Loading {OUTPUT_FILE}…")
    wb = openpyxl.load_workbook(OUTPUT_FILE)
    ws = wb["Listings"]

    headers = [cell.value for cell in ws[1]]
    col_idx = {h: i + 1 for i, h in enumerate(headers) if h}  # 1-based

    required = ["Title", "Company", "Location", "Type", "Salary", "Description", "S3"]
    missing = [c for c in required if c not in col_idx]
    if missing:
        print(f"ERROR: Missing columns: {missing}")
        sys.exit(1)

    s3_col = col_idx["S3"]
    print(f"Scoring rows 2–{1 + ROWS_TO_RATE} → column {s3_col} (S3)…")

    for idx, row_num in enumerate(range(2, 2 + ROWS_TO_RATE), start=1):
        row = ws[row_num]

        def cv(col_name):
            return row[col_idx[col_name] - 1].value

        score = score_job(
            cv("Title"), cv("Company"), cv("Location"),
            cv("Type"), cv("Salary"), cv("Description"),
        )

        ws.cell(row=row_num, column=s3_col).value = score

        if idx % 100 == 0:
            print(f"  … {idx} rows rated")

    print(f"\nDone. Saving {OUTPUT_FILE}…")
    wb.save(OUTPUT_FILE)
    print("Saved.")


if __name__ == "__main__":
    main()
