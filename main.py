#!/usr/bin/env python3
"""
UK PM Job Scraper -- main entry point.
"""
import argparse
import time
from datetime import datetime

from scraper_reed import scrape_reed
from scraper_cvlibrary import scrape_cvlibrary
from scraper_totaljobs import scrape_totaljobs
from scraper_linkedin import scrape_linkedin, scrape_linkedin_extended
from output import write_listings, get_existing_urls, rescore_file, rescore_s2_file, get_incomplete_rows, write_refetched, recolor_by_date
from models import (
    deduplicate, score_listings, check_onsite_days, score_job,
    check_description_filter,
    JobListing, ExcludedJob,
)
import config


SCRAPERS = {
    "linkedin": ("LinkedIn", scrape_linkedin),
    "linkedin-extended": ("LinkedIn (Extended)", scrape_linkedin_extended),
    "reed": ("Reed", scrape_reed),
    "cvlibrary": ("CV-Library", scrape_cvlibrary),
    "totaljobs": ("TotalJobs", scrape_totaljobs),
}


def run_scrapers(sources=None, known_urls: set[str] = None) -> tuple[list[JobListing], list[ExcludedJob]]:
    all_listings = []
    all_excluded = []

    if sources is None:
        sources = list(SCRAPERS.keys())

    for key in sources:
        name, func = SCRAPERS[key]
        t0 = time.time()
        jobs, excluded = func(known_urls=known_urls)
        elapsed = time.time() - t0
        mins, secs = divmod(int(elapsed), 60)
        print(f"[{name}] Completed in {mins}m {secs}s")
        all_listings.extend(jobs)
        all_excluded.extend(excluded)

    return all_listings, all_excluded


# ---------------------------------------------------------------------------
# Refetch mode: re-fetch missing descriptions for existing spreadsheet rows
# ---------------------------------------------------------------------------

# Map source names to the detail-fetching functions.
# Each returns a dict: {description, salary, date_posted, ...}

def _refetch_reed(url: str) -> dict:
    """Fetch description for a Reed job URL."""
    import re
    from scraper_reed import get_job_details
    from models import clean_html
    # Extract job ID from URL: reed.co.uk/jobs/TITLE/JOBID
    match = re.search(r'/jobs?/.*?(\d{5,})', url)
    if not match:
        match = re.search(r'(\d{5,})', url)
    if not match:
        return {}
    job_id = int(match.group(1))
    try:
        details = get_job_details(job_id)
        return {"description": clean_html(details.get("jobDescription", ""))}
    except Exception as e:
        print(f"[Refetch]   -> Reed error for {url}: {e}")
        return {}


def _refetch_with_selenium(driver, url: str, source: str) -> dict:
    """Generic Selenium detail fetch — works for CV-Library, TotalJobs, LinkedIn."""
    import re
    from selenium.common.exceptions import WebDriverException

    if source == "CV-Library":
        from scraper_cvlibrary import _get_description_selenium, _dismiss_cookies
        try:
            desc = _get_description_selenium(driver, url)
            return {"description": desc}
        except Exception as e:
            print(f"[Refetch]   -> CV-Library error: {e}")
            return {}

    elif source == "TotalJobs":
        from scraper_totaljobs import _get_description
        try:
            return _get_description(driver, url)
        except Exception as e:
            print(f"[Refetch]   -> TotalJobs error: {e}")
            return {}

    elif source == "LinkedIn":
        from scraper_linkedin import _extract_detail_from_tab
        try:
            driver.get(url)
            import time
            time.sleep(3)
            return _extract_detail_from_tab(driver)
        except Exception as e:
            print(f"[Refetch]   -> LinkedIn error: {e}")
            return {}

    return {}


def _run_refetch(filepath: str):
    """Find incomplete rows, re-fetch descriptions, score, and write back."""
    import time as _time

    incomplete = get_incomplete_rows(filepath)
    if not incomplete:
        print("[Refetch] All rows have descriptions and scores. Nothing to do.")
        return

    # Separate: rows needing description vs only needing score
    needs_desc = [r for r in incomplete if not r["has_desc"]]
    needs_score_only = [r for r in incomplete if r["has_desc"] and not r["has_score"]]

    print(f"[Refetch] Found {len(incomplete)} incomplete rows:")
    print(f"  Missing description: {len(needs_desc)}")
    print(f"  Missing score only:  {len(needs_score_only)}")

    # Group needs_desc by source
    by_source = {}
    for row in needs_desc:
        src = row["source"]
        by_source.setdefault(src, []).append(row)

    for src, rows in sorted(by_source.items()):
        print(f"  {src}: {len(rows)} to refetch")

    updates = []

    # --- Handle score-only rows first (no fetching needed) ---
    if needs_score_only:
        print(f"\n[Refetch] Scoring {len(needs_score_only)} rows that have descriptions but no score...")
        # We need to read descriptions from the file
        from openpyxl import load_workbook as _lwb
        wb = _lwb(filepath)
        for row_info in needs_score_only:
            ws = wb[row_info["sheet_name"]]
            from output import _read_headers, _col_map
            headers = _read_headers(ws)
            cmap = _col_map(headers)
            desc_col = cmap.get("Description")
            loc_col = cmap.get("Location")
            type_col = cmap.get("Type")
            sal_col = cmap.get("Salary")
            if desc_col:
                desc = ws.cell(row=row_info["row_num"], column=desc_col).value or ""
                loc = ws.cell(row=row_info["row_num"], column=loc_col).value or "" if loc_col else ""
                wt = ws.cell(row=row_info["row_num"], column=type_col).value or "" if type_col else ""
                sal = ws.cell(row=row_info["row_num"], column=sal_col).value or "" if sal_col else ""

                job = JobListing(
                    source=row_info["source"], title=row_info["title"],
                    company="", location=loc, salary=sal,
                    url=row_info["url"], description=desc,
                    date_posted="", work_type=wt,
                )
                score = score_job(job)
                updates.append({
                    "sheet_name": row_info["sheet_name"],
                    "row_num": row_info["row_num"],
                    "score": score,
                })
        wb.close()
        print(f"[Refetch] Scored {len(needs_score_only)} rows")

    # --- Fetch missing descriptions ---
    if needs_desc:
        # Reed: use API (no browser needed)
        reed_rows = by_source.get("Reed", [])
        if reed_rows:
            print(f"\n[Refetch] Fetching {len(reed_rows)} Reed descriptions (API)...")
            for i, row_info in enumerate(reed_rows):
                detail = _refetch_reed(row_info["url"])
                desc = detail.get("description", "")
                if desc:
                    # Filter + score
                    passes, reason = check_description_filter(desc)
                    if passes:
                        from models import detect_work_type
                        job = JobListing(
                            source="Reed", title=row_info["title"],
                            company="", location="", salary="",
                            url=row_info["url"], description=desc,
                            date_posted="", work_type="",
                        )
                        score = score_job(job)
                        updates.append({
                            "sheet_name": row_info["sheet_name"],
                            "row_num": row_info["row_num"],
                            "description": desc,
                            "score": score,
                        })
                        print(f"[Refetch]   -> {row_info['title'][:50]}: fetched, score={score}")
                    else:
                        print(f"[Refetch]   -> {row_info['title'][:50]}: excluded ({reason})")
                else:
                    print(f"[Refetch]   -> {row_info['title'][:50]}: no description found")
                _time.sleep(0.3)

        # Selenium-based sources: CV-Library, TotalJobs
        for source_name in ["CV-Library", "TotalJobs"]:
            source_rows = by_source.get(source_name, [])
            if not source_rows:
                continue

            print(f"\n[Refetch] Fetching {len(source_rows)} {source_name} descriptions (Selenium)...")

            if source_name == "CV-Library":
                from scraper_cvlibrary import _make_driver, _is_session_alive, _restart_driver
            elif source_name == "TotalJobs":
                from scraper_totaljobs import _make_driver, _is_session_alive, _restart_driver

            try:
                driver = _make_driver()
            except Exception as e:
                print(f"[Refetch]   -> Could not start browser for {source_name}: {e}")
                continue

            try:
                for i, row_info in enumerate(source_rows):
                    if not _is_session_alive(driver):
                        driver = _restart_driver(driver)
                        if driver is None:
                            print(f"[Refetch]   -> Browser dead, stopping {source_name}")
                            break

                    detail = _refetch_with_selenium(driver, row_info["url"], source_name)
                    desc = detail.get("description", "")
                    if desc:
                        passes, reason = check_description_filter(desc)
                        if passes:
                            job = JobListing(
                                source=source_name, title=row_info["title"],
                                company="", location="", salary="",
                                url=row_info["url"], description=desc,
                                date_posted="", work_type="",
                            )
                            score = score_job(job)
                            updates.append({
                                "sheet_name": row_info["sheet_name"],
                                "row_num": row_info["row_num"],
                                "description": desc,
                                "score": score,
                            })
                            print(f"[Refetch]   -> {row_info['title'][:50]}: fetched, score={score}")
                        else:
                            print(f"[Refetch]   -> {row_info['title'][:50]}: excluded ({reason})")
                    else:
                        print(f"[Refetch]   -> {row_info['title'][:50]}: no description found")
                    _time.sleep(0.5)
            finally:
                try:
                    driver.quit()
                except Exception:
                    pass

        # LinkedIn: uses subprocess Chrome
        linkedin_rows = by_source.get("LinkedIn", [])
        if linkedin_rows:
            print(f"\n[Refetch] Fetching {len(linkedin_rows)} LinkedIn descriptions...")
            from scraper_linkedin import _make_driver as _make_li_driver
            from scraper_linkedin import _is_session_alive as _li_alive
            from scraper_linkedin import _check_login

            try:
                driver = _make_li_driver()
            except Exception as e:
                print(f"[Refetch]   -> Could not start LinkedIn browser: {e}")
                linkedin_rows = []

            if linkedin_rows:
                try:
                    if not _check_login(driver):
                        print("[Refetch]   -> Not logged into LinkedIn, skipping")
                        linkedin_rows = []

                    for i, row_info in enumerate(linkedin_rows):
                        if not _li_alive(driver):
                            print("[Refetch]   -> LinkedIn browser dead, stopping")
                            break

                        detail = _refetch_with_selenium(driver, row_info["url"], "LinkedIn")
                        desc = detail.get("description", "")
                        if desc:
                            passes, reason = check_description_filter(desc)
                            if passes:
                                job = JobListing(
                                    source="LinkedIn", title=row_info["title"],
                                    company="", location="", salary="",
                                    url=row_info["url"], description=desc,
                                    date_posted="", work_type="",
                                )
                                score = score_job(job)
                                updates.append({
                                    "sheet_name": row_info["sheet_name"],
                                    "row_num": row_info["row_num"],
                                    "description": desc,
                                    "score": score,
                                })
                                print(f"[Refetch]   -> {row_info['title'][:50]}: fetched, score={score}")
                            else:
                                print(f"[Refetch]   -> {row_info['title'][:50]}: excluded ({reason})")
                        else:
                            print(f"[Refetch]   -> {row_info['title'][:50]}: no description found")
                        _time.sleep(2)
                finally:
                    try:
                        proc = getattr(driver, "_chrome_proc", None)
                        driver.quit()
                        if proc:
                            proc.terminate()
                    except Exception:
                        pass

    # Write all updates back
    if updates:
        print(f"\n[Refetch] Writing {len(updates)} updates to spreadsheet...")
        write_refetched(filepath, updates)
    else:
        print("\n[Refetch] No updates to write.")


def main():
    parser = argparse.ArgumentParser(description="UK PM Job Scraper")
    parser.add_argument("--with-summaries", action="store_true")
    parser.add_argument("--output", default=config.OUTPUT_FILE)
    parser.add_argument("--rescore", action="store_true",
                        help="Re-score all jobs in the spreadsheet without scraping (updates S1 + S2).")
    parser.add_argument("--rescore-s2", action="store_true",
                        help="Re-score only S2 (CV-fit) for all jobs without scraping. Faster than --rescore.")
    parser.add_argument("--refetch", action="store_true",
                        help="Re-fetch missing descriptions/scores for existing listings.")
    parser.add_argument("--recolor", action="store_true",
                        help="Recolor company cells by scraping date (latest=bright yellow, prev=pale yellow).")
    parser.add_argument(
        "sources", nargs="*",
        help=f"Scraper(s) to run: {', '.join(SCRAPERS.keys())}. Omit for all.",
    )
    args = parser.parse_args()

    # Rescore mode: just score and exit
    if args.rescore:
        print(f"{'='*60}")
        print(f"UK PM Job Scraper -- Rescore mode (S1 + S2)")
        print(f"{'='*60}\n")
        rescore_file(args.output)
        return

    # Rescore S2 only (CV-fit): faster, skips S1
    if getattr(args, "rescore_s2", False):
        print(f"{'='*60}")
        print(f"UK PM Job Scraper -- Rescore S2 mode (CV-fit only)")
        print(f"{'='*60}\n")
        rescore_s2_file(args.output)
        return

    # Recolor mode: recolor company cells by scraping date
    if args.recolor:
        print(f"{'='*60}")
        print(f"UK PM Job Scraper -- Recolor mode")
        print(f"{'='*60}\n")
        recolor_by_date(args.output)
        return

    # Refetch mode: re-fetch missing descriptions, then score and write back
    if args.refetch:
        print(f"{'='*60}")
        print(f"UK PM Job Scraper -- Refetch mode")
        print(f"{'='*60}\n")
        _run_refetch(args.output)
        return

    # Separate positional args: .xlsx files are output, rest are sources
    source_args = []
    for s in (args.sources or []):
        if s.lower().endswith(".xlsx"):
            args.output = s
        else:
            source_args.append(s)

    # Validate source names
    if source_args:
        for s in source_args:
            if s.lower() not in SCRAPERS:
                parser.error(f"Unknown source '{s}'. Choose from: {', '.join(SCRAPERS.keys())}")
        sources = [s.lower() for s in source_args]
        label = ", ".join(SCRAPERS[s][0] for s in sources)
        is_full_run = False
    else:
        sources = None
        label = "all"
        is_full_run = True

    print(f"{'='*60}")
    print(f"UK PM Job Scraper -- {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print(f"Sources: {label}")
    if args.output != config.OUTPUT_FILE:
        print(f"Output:  {args.output}")
    print(f"{'='*60}\n")

    start_time = time.time()

    # 1. Load existing URLs from spreadsheet (used to skip detail fetches)
    existing_urls = get_existing_urls(args.output)
    print(f"Existing URLs in spreadsheet: {len(existing_urls)}\n")

    # 2. Scrape
    all_listings, all_excluded = run_scrapers(sources, known_urls=existing_urls)
    print(f"\n{'-'*40}")
    print(f"Total accepted from all sources: {len(all_listings)}")
    print(f"Total excluded from all sources: {len(all_excluded)}")

    # 3. Deduplicate against existing spreadsheet + 3-stage internal dedup
    print(f"Existing URLs in spreadsheet: {len(existing_urls)}")
    new_listings = deduplicate(all_listings, existing_urls)
    print(f"New unique listings to add: {len(new_listings)}")

    # 3. Onsite days filter (disabled — keeping all jobs regardless of onsite requirement)
    # filtered_listings = []
    # for job in new_listings:
    #     if job.description:
    #         passes, reason = check_onsite_days(job.description)
    #         if not passes:
    #             all_excluded.append(ExcludedJob(
    #                 job.source, job.title, job.company,
    #                 job.location, job.url, reason,
    #             ))
    #             continue
    #     filtered_listings.append(job)
    #
    # onsite_rejected = len(new_listings) - len(filtered_listings)
    # if onsite_rejected:
    #     print(f"[Filter] Rejected {onsite_rejected} listings for onsite days requirement")
    # new_listings = filtered_listings

    # Deduplicate excluded list by URL
    seen_excluded = set()
    unique_excluded = []
    for ex in all_excluded:
        key = ex.url.lower().strip()
        if key not in seen_excluded:
            seen_excluded.add(key)
            unique_excluded.append(ex)
    all_excluded = unique_excluded

    if not new_listings and not all_excluded:
        print("\nNo new listings to add. Done.")
        return

    # 4. Score all new listings
    if new_listings:
        print()
        score_listings(new_listings)

    # 5. Split into main (score >= 60 or unscored) and low score (< 60)
    main_listings = []
    low_score_listings = []
    for job in new_listings:
        if job.initial_score is not None and job.initial_score < 60:
            low_score_listings.append(job)
        else:
            main_listings.append(job)

    if low_score_listings:
        print(f"[Split] {len(main_listings)} to main sheet, {len(low_score_listings)} to low-score sheet")

    # 6. Write to spreadsheet
    print()
    write_listings(
        main_listings,
        low_score_listings=low_score_listings,
        excluded=all_excluded,
        filepath=args.output,
        recolor_existing=is_full_run,
    )

    # 7. Stats
    print(f"\n{'='*60}")
    print("Summary:")
    sources = {}
    for job in new_listings:
        sources[job.source] = sources.get(job.source, 0) + 1
    for source, count in sorted(sources.items()):
        print(f"  {source}: {count} new listings")
    print(f"  Main sheet: {len(main_listings)}")
    print(f"  Low score:  {len(low_score_listings)}")
    print(f"  Excluded:   {len(all_excluded)}")
    elapsed = time.time() - start_time
    mins, secs = divmod(int(elapsed), 60)
    print(f"  Time:       {mins}m {secs}s")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
