"""
Spreadsheet output: append accepted listings, rewrite excluded sheet each run.
URLs are clickable hyperlinks. S1 column is color-coded.
New listings get bright yellow Company cells; previous run's yellow demotes
to pale yellow; older listings go white. 3-tier highlight system.
Sub-60 score jobs go to a separate "Low Score" sheet.

IMPORTANT: When appending to an existing file, column positions are read
from the file's header row â€” so custom column orders and extra columns
(S2, S3, Applied?, etc.) are preserved and never overwritten.
"""
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import config
from models import JobListing, ExcludedJob, SPREADSHEET_COLUMNS, EXCLUDED_COLUMNS, score_color


HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
EXCLUDED_HEADER_FILL = PatternFill("solid", fgColor="943634")
LOW_SCORE_HEADER_FILL = PatternFill("solid", fgColor="7F6000")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

BODY_FONT = Font(name="Arial", size=10)
URL_FONT = Font(name="Arial", size=10, color="0563C1", underline="single")
BODY_ALIGN = Alignment(vertical="top", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

LISTINGS_COL_WIDTHS = {
    "Date Scraped": 12, "Date Posted": 12, "Source": 10, "Title": 40,
    "Company": 25, "Location": 20, "Type": 12, "Salary": 22,
    "S1": 8, "URL": 50, "Description": 80,
}

EXCLUDED_COL_WIDTHS = {
    "A": 10, "B": 40, "C": 25, "D": 20, "E": 50, "F": 40, "G": 40,
}

EVEN_ROW_FILL = PatternFill("solid", fgColor="F2F2F2")
NEW_COMPANY_FILL = PatternFill("solid", fgColor="FFFF00")       # bright yellow: newest run
PREV_COMPANY_FILL = PatternFill("solid", fgColor="FFFFCC")      # pale yellow: previous run
NO_FILL = PatternFill(fill_type=None)

# The fields we write, mapped from header name -> JobListing attribute
# This defines what the scraper owns. Other columns (S2, S3, Applied?, etc.) are left alone.
FIELD_MAP = {
    "Date Scraped": lambda j: j.date_scraped,
    "Date Posted": lambda j: j.date_posted,
    "Source": lambda j: j.source,
    "Title": lambda j: j.title,
    "Company": lambda j: j.company,
    "Location": lambda j: j.location,
    "Type": lambda j: j.work_type,
    "Salary": lambda j: j.salary,
    "S1": lambda j: j.initial_score if j.initial_score is not None else "",
    "S2": lambda j: j.secondary_score if j.secondary_score is not None else "",
    "URL": lambda j: j.url,
    "Description": lambda j: j.description,
    "Duplicates": lambda j: "\n".join(j.duplicate_urls) if j.duplicate_urls else "",
}

# Default column order for new files
DEFAULT_COLUMNS = [
    "Date Scraped", "Date Posted", "Source", "Title", "Company",
    "Location", "Type", "Salary", "URL", "Description", "S1", "S2", "Duplicates",
]


def _read_headers(ws) -> list[str]:
    """Read header names from row 1."""
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(str(cell.value))
        else:
            break
    return headers


def _col_map(headers: list[str]) -> dict[str, int]:
    """Map header name -> 1-based column index."""
    return {h: i + 1 for i, h in enumerate(headers)}


def _apply_header(ws, columns, header_fill=HEADER_FILL):
    # Default row height: 40px â‰ˆ 30pt
    ws.sheet_format.defaultRowHeight = 30
    for col_idx, header in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = header_fill
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"
    ws.freeze_panes = "A2"
    # Apply widths by name
    for col_idx, header in enumerate(columns, 1):
        if header in LISTINGS_COL_WIDTHS:
            ws.column_dimensions[get_column_letter(col_idx)].width = LISTINGS_COL_WIDTHS[header]


def _apply_excluded_header(ws, columns, header_fill=EXCLUDED_HEADER_FILL):
    for col_idx, header in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = header_fill
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"
    ws.freeze_panes = "A2"
    for letter, width in EXCLUDED_COL_WIDTHS.items():
        ws.column_dimensions[letter].width = width


def _style_cell(cell, row_num, is_new=False, is_company=False):
    """Style a single data cell."""
    cell.font = BODY_FONT
    cell.alignment = BODY_ALIGN
    cell.border = THIN_BORDER
    if is_company:
        cell.fill = NEW_COMPANY_FILL if is_new else NO_FILL
    elif row_num % 2 == 0:
        cell.fill = EVEN_ROW_FILL


def _write_url(ws, row: int, col: int, url: str):
    """Write a clickable hyperlink."""
    cell = ws.cell(row=row, column=col)
    if url and str(url).startswith("http"):
        cell.hyperlink = url
        cell.value = url
        cell.font = URL_FONT
    else:
        cell.value = url
    cell.border = THIN_BORDER
    cell.alignment = BODY_ALIGN


def _write_score(ws, row: int, col: int, score_value):
    """Write score with color coding."""
    cell = ws.cell(row=row, column=col)
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center")

    if score_value is not None and score_value != "":
        score_int = int(score_value)
        cell.value = score_int
        fill_hex, font_hex = score_color(score_int)
        cell.fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
        cell.font = Font(bold=True, size=11, name="Arial", color=font_hex)
    else:
        cell.value = "N/A"
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.font = Font(italic=True, color="888888", name="Arial")


def get_existing_urls(filepath: str) -> set[str]:
    p = Path(filepath)
    if not p.exists():
        return set()
    wb = load_workbook(filepath, read_only=True)
    urls = set()
    for sheet_name in [config.SHEET_NAME, config.LOW_SCORE_SHEET_NAME]:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            if "URL" in header_row:
                url_col = header_row.index("URL") + 1
                for row in ws.iter_rows(min_row=2, min_col=url_col, max_col=url_col, values_only=True):
                    if row[0]:
                        urls.add(str(row[0]).strip().lower())
    wb.close()
    return urls


def _demote_company_highlights(ws, company_col: int):
    """
    3-tier highlight demotion for company cells:
      bright yellow (FFFF00, newest) → pale yellow (FFFFCC, previous) → white/no fill
    Called before writing new rows so the tiers shift down each run.
    """
    if company_col == 0:
        return
    for row_num in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_num, column=company_col)
        fill_color = ""
        if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
            fill_color = str(cell.fill.fgColor.rgb).upper()
            # openpyxl returns ARGB (8 chars) after save/reload — strip alpha prefix
            if len(fill_color) == 8:
                fill_color = fill_color[2:]

        if fill_color == "FFFF00":
            # Bright yellow → pale yellow
            cell.fill = PREV_COMPANY_FILL
        elif fill_color == "FFFFCC":
            # Pale yellow → no fill (respect even/odd row striping)
            cell.fill = EVEN_ROW_FILL if row_num % 2 == 0 else NO_FILL
        # Everything else stays as-is


def _write_job_rows(ws, listings, start_row, cmap, is_new=False):
    """
    Write job listing rows using the column map from the sheet's headers.
    Only writes to columns that exist in both FIELD_MAP and the sheet.
    Returns number of rows written.
    """
    url_col = cmap.get("URL", 0)
    score_col = cmap.get("S1", 0)
    company_col = cmap.get("Company", 0)

    added = 0
    row_num = start_row
    for job in listings:
        for header, getter in FIELD_MAP.items():
            col = cmap.get(header)
            if col is None:
                continue
            value = getter(job)

            if header == "URL":
                _write_url(ws, row_num, col, value)
            elif header in ("S1", "S2"):
                _write_score(ws, row_num, col, value)
            else:
                cell = ws.cell(row=row_num, column=col, value=value)
                _style_cell(cell, row_num, is_new=is_new, is_company=(header == "Company"))

        row_num += 1
        added += 1
    return added


def write_listings(
    listings: list[JobListing],
    low_score_listings: list[JobListing] = None,
    excluded: list[ExcludedJob] = None,
    filepath: str = None,
    recolor_existing: bool = True,
):
    filepath = filepath or config.OUTPUT_FILE
    p = Path(filepath)

    if p.exists():
        wb = load_workbook(filepath)
        if config.SHEET_NAME in wb.sheetnames:
            ws = wb[config.SHEET_NAME]
            ws.sheet_format.defaultRowHeight = 30
            headers = _read_headers(ws)
            cmap = _col_map(headers)
        else:
            ws = wb.create_sheet(config.SHEET_NAME)
            _apply_header(ws, DEFAULT_COLUMNS)
            headers = DEFAULT_COLUMNS
            cmap = _col_map(headers)
        # Demote highlight tiers: yellow → pale yellow → white (full run only)
        if recolor_existing:
            _demote_company_highlights(ws, cmap.get("Company", 0))
        next_row = ws.max_row + 1
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = config.SHEET_NAME
        _apply_header(ws, DEFAULT_COLUMNS)
        headers = DEFAULT_COLUMNS
        cmap = _col_map(headers)
        next_row = 2

    # Warn about missing columns
    missing = [h for h in FIELD_MAP if h not in cmap]
    if missing:
        print(f"[Output] WARNING: columns missing from sheet, will skip: {missing}")

    # Write accepted listings with yellow company
    added = _write_job_rows(ws, listings, next_row, cmap, is_new=True)

    # --- Low Score sheet (sub-60) -- wipe and rewrite every run ---
    if low_score_listings is not None:
        if config.LOW_SCORE_SHEET_NAME in wb.sheetnames:
            del wb[config.LOW_SCORE_SHEET_NAME]
        ws_low = wb.create_sheet(config.LOW_SCORE_SHEET_NAME)
        _apply_header(ws_low, DEFAULT_COLUMNS, LOW_SCORE_HEADER_FILL)
        low_cmap = _col_map(DEFAULT_COLUMNS)
        _write_job_rows(ws_low, low_score_listings, 2, low_cmap, is_new=False)
        print(f"[Output] Wrote {len(low_score_listings)} low-score listings to '{config.LOW_SCORE_SHEET_NAME}' sheet")

    # --- Excluded sheet -- wipe and rewrite every run ---
    if excluded is not None:
        if config.EXCLUDED_SHEET_NAME in wb.sheetnames:
            del wb[config.EXCLUDED_SHEET_NAME]
        ws_ex = wb.create_sheet(config.EXCLUDED_SHEET_NAME)
        _apply_excluded_header(ws_ex, EXCLUDED_COLUMNS)

        ex_url_col = EXCLUDED_COLUMNS.index("URL") + 1
        for i, ex in enumerate(excluded):
            row_num = i + 2
            row_data = ex.to_row()
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_ex.cell(row=row_num, column=col_idx, value=value)
                if col_idx == ex_url_col:
                    _write_url(ws_ex, row_num, col_idx, value)
                else:
                    _style_cell(cell, row_num)

        print(f"[Output] Wrote {len(excluded)} excluded listings to '{config.EXCLUDED_SHEET_NAME}' sheet")

    wb.save(filepath)
    total_main = (next_row - 2) + added
    print(f"[Output] Wrote {added} new listings to '{config.SHEET_NAME}' sheet (total rows: {total_main})")
    return added


def _rescore_sheet(ws, cmap, score_s1: bool, score_s2: bool):
    """
    Inner helper: re-score all rows in one worksheet.
    Writes S1 and/or S2 based on flags.  Returns count of rows processed.
    """
    from models import JobListing, score_job, score_job_s2

    title_col   = cmap.get("Title")
    desc_col    = cmap.get("Description")
    company_col = cmap.get("Company")
    location_col = cmap.get("Location")
    type_col    = cmap.get("Type")
    salary_col  = cmap.get("Salary")
    source_col  = cmap.get("Source")
    url_col     = cmap.get("URL")
    s1_col      = cmap.get("S1")
    s2_col      = cmap.get("S2")

    if not title_col or not desc_col:
        return 0

    count = 0
    for row_num in range(2, ws.max_row + 1):
        title = ws.cell(row=row_num, column=title_col).value or ""
        if not title:
            continue

        desc = ws.cell(row=row_num, column=desc_col).value or ""
        job = JobListing(
            source=ws.cell(row=row_num, column=source_col).value or "" if source_col else "",
            title=title,
            company=ws.cell(row=row_num, column=company_col).value or "" if company_col else "",
            location=ws.cell(row=row_num, column=location_col).value or "" if location_col else "",
            salary=ws.cell(row=row_num, column=salary_col).value or "" if salary_col else "",
            url=ws.cell(row=row_num, column=url_col).value or "" if url_col else "",
            description=desc,
            date_posted="",
            work_type=ws.cell(row=row_num, column=type_col).value or "" if type_col else "",
        )

        if score_s1 and s1_col:
            _write_score(ws, row_num, s1_col, score_job(job))
        if score_s2 and s2_col:
            _write_score(ws, row_num, s2_col, score_job_s2(job))

        count += 1
    return count


def rescore_file(filepath: str = None):
    """
    Re-score all jobs in the spreadsheet (both S1 and S2) and write back in-place.
    Works on Listings and Low Score sheets.
    """
    filepath = filepath or config.OUTPUT_FILE
    p = Path(filepath)
    if not p.exists():
        print(f"[Rescore] File not found: {filepath}")
        return

    wb = load_workbook(filepath)
    total = 0

    for sheet_name in [config.SHEET_NAME, config.LOW_SCORE_SHEET_NAME]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        cmap = _col_map(_read_headers(ws))
        if "S1" not in cmap:
            print(f"[Rescore] No 'S1' column in '{sheet_name}', skipping")
            continue
        n = _rescore_sheet(ws, cmap, score_s1=True, score_s2=("S2" in cmap))
        total += n
        s2_note = " + S2" if "S2" in cmap else ""
        print(f"[Rescore] Re-scored {n} jobs (S1{s2_note}) in '{sheet_name}'")

    wb.save(filepath)
    print(f"[Rescore] Done. {total} jobs re-scored in {filepath}")


def rescore_s2_file(filepath: str = None):
    """
    Re-score only S2 (CV-fit) for all jobs in the spreadsheet in-place.
    Faster than --rescore because it skips S1.  Works on Listings and Low Score sheets.
    """
    filepath = filepath or config.OUTPUT_FILE
    p = Path(filepath)
    if not p.exists():
        print(f"[Rescore S2] File not found: {filepath}")
        return

    wb = load_workbook(filepath)
    total = 0

    for sheet_name in [config.SHEET_NAME, config.LOW_SCORE_SHEET_NAME]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        cmap = _col_map(_read_headers(ws))
        if "S2" not in cmap:
            print(f"[Rescore S2] No 'S2' column in '{sheet_name}', skipping")
            continue
        n = _rescore_sheet(ws, cmap, score_s1=False, score_s2=True)
        total += n
        print(f"[Rescore S2] Re-scored {n} jobs in '{sheet_name}'")

    wb.save(filepath)
    print(f"[Rescore S2] Done. {total} jobs re-scored in {filepath}")


def get_incomplete_rows(filepath: str = None) -> list[dict]:
    """
    Find rows in Listings/Low Score sheets that have a URL but are missing
    Description and/or S1 score.
    Returns list of dicts with: sheet_name, row_num, source, url, title, has_desc, has_score.
    """
    filepath = filepath or config.OUTPUT_FILE
    p = Path(filepath)
    if not p.exists():
        return []

    wb = load_workbook(filepath)
    incomplete = []

    for sheet_name in [config.SHEET_NAME, config.LOW_SCORE_SHEET_NAME]:
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        headers = _read_headers(ws)
        cmap = _col_map(headers)

        url_col = cmap.get("URL")
        desc_col = cmap.get("Description")
        s1_col = cmap.get("S1")
        source_col = cmap.get("Source")
        title_col = cmap.get("Title")

        if not url_col:
            continue

        for row_num in range(2, ws.max_row + 1):
            url = ws.cell(row=row_num, column=url_col).value
            if not url or not str(url).startswith("http"):
                continue

            desc = ws.cell(row=row_num, column=desc_col).value if desc_col else None
            score = ws.cell(row=row_num, column=s1_col).value if s1_col else None

            has_desc = bool(desc and str(desc).strip() and len(str(desc).strip()) > 20)
            has_score = (score is not None and score != "" and score != "N/A")

            if not has_desc or not has_score:
                source = ws.cell(row=row_num, column=source_col).value if source_col else ""
                title = ws.cell(row=row_num, column=title_col).value if title_col else ""
                incomplete.append({
                    "sheet_name": sheet_name,
                    "row_num": row_num,
                    "source": str(source or "").strip(),
                    "url": str(url).strip(),
                    "title": str(title or "").strip(),
                    "has_desc": has_desc,
                    "has_score": has_score,
                })

    wb.close()
    return incomplete


def write_refetched(filepath: str, updates: list[dict]):
    """
    Write re-fetched descriptions and scores back to the spreadsheet in-place.
    Each update dict: {sheet_name, row_num, description, score}.
    """
    from models import score_color

    filepath = filepath or config.OUTPUT_FILE
    p = Path(filepath)
    if not p.exists():
        print(f"[Refetch] File not found: {filepath}")
        return

    wb = load_workbook(filepath)
    written = 0

    for upd in updates:
        sheet_name = upd["sheet_name"]
        row_num = upd["row_num"]

        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        headers = _read_headers(ws)
        cmap = _col_map(headers)

        desc_col = cmap.get("Description")
        s1_col = cmap.get("S1")

        if upd.get("description") and desc_col:
            cell = ws.cell(row=row_num, column=desc_col, value=upd["description"])
            _style_cell(cell, row_num)

        if upd.get("score") is not None and s1_col:
            _write_score(ws, row_num, s1_col, upd["score"])

        written += 1

    wb.save(filepath)
    print(f"[Refetch] Updated {written} rows in {filepath}")


def recolor_by_date(filepath: str = None):
    """
    Recolor company cells in the Listings sheet based on Date Scraped.
    Latest scraping date → bright yellow, previous → pale yellow, older → no fill.
    """
    filepath = filepath or config.OUTPUT_FILE
    p = Path(filepath)
    if not p.exists():
        print(f"[Recolor] File not found: {filepath}")
        return

    wb = load_workbook(filepath)
    if config.SHEET_NAME not in wb.sheetnames:
        print(f"[Recolor] Sheet '{config.SHEET_NAME}' not found")
        wb.close()
        return

    ws = wb[config.SHEET_NAME]
    headers = _read_headers(ws)
    cmap = _col_map(headers)

    date_col = cmap.get("Date Scraped")
    company_col = cmap.get("Company")

    if not date_col or not company_col:
        print("[Recolor] Missing 'Date Scraped' or 'Company' column")
        wb.close()
        return

    # Collect date per row
    row_dates: dict[int, str] = {}
    for row_num in range(2, ws.max_row + 1):
        val = ws.cell(row=row_num, column=date_col).value
        if val:
            row_dates[row_num] = str(val)

    unique_dates = sorted(set(row_dates.values()), reverse=True)
    if not unique_dates:
        print("[Recolor] No dates found, nothing to do")
        wb.close()
        return

    latest = unique_dates[0]
    second = unique_dates[1] if len(unique_dates) > 1 else None

    print(f"[Recolor] {len(unique_dates)} scraping dates found:")
    print(f"  {latest} → bright yellow (latest)")
    if second:
        print(f"  {second} → pale yellow (previous)")
    if len(unique_dates) > 2:
        print(f"  {len(unique_dates) - 2} older date(s) → no color")

    for row_num, date_str in row_dates.items():
        cell = ws.cell(row=row_num, column=company_col)
        if date_str == latest:
            cell.fill = NEW_COMPANY_FILL
        elif date_str == second:
            cell.fill = PREV_COMPANY_FILL
        else:
            cell.fill = EVEN_ROW_FILL if row_num % 2 == 0 else NO_FILL

    wb.save(filepath)
    print(f"[Recolor] Updated {len(row_dates)} company cells in '{filepath}'")
